import boto3
import datetime
import openpyxl


workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Cloudtrail Events'

worksheet.cell(row=1, column=1, value='Instance ID')
worksheet.cell(row=1, column=2, value='Instance Name')
worksheet.cell(row=1, column=3, value='EventName')
worksheet.cell(row=1, column=4, value='Username')
worksheet.cell(row=1, column=5, value='AccesKeyID')
worksheet.cell(row=1, column=6, value='EventTime')

row_num=2
endtime =datetime.datetime.now()
interval = datetime.timedelta(days=7)
starttime = endtime - interval

cloudtrail = boto3.client('cloudtrail')
ec2 = boto3.client('ec2')
paginator = cloudtrail.get_paginator('lookup_events')
    page_iterator = paginator.paginate(
    LookupAttributes=[
    {
        'AttributeKey': 'ResourceType',
        'AttributeValue': 'AWS::EC2::Instance'}
    ],
    StartTime=starttime,
    EndTime=endtime
    )

for page in page_iterator:
    for event in page['Events']:
        resources = event['Resources']
        for resource in resources:
            if resource['ResourceType'] == 'AWS::EC2::Instance':
                instance_id = resource['ResourceName']
                try:
                    instances = ec2.describe_instances(InstanceIds=[instance_id])
                    for reservation in instances['Reservations']:
                        for instance in reservation['Instances']:
                            for tag in instance['Tags']:
                                if tag['Key'] == 'Name':
                                    instanceName= tag['Value']
                except Exception as e:
                    instanceName = "Instance doesn't exist"            
        eventname = event['EventName']
        username  = event['Username']
        try:
            access_key = event['AccessKeyId']
        except KeyError:
            access_key ="N.A"
        eventtime = event['EventTime']

        worksheet.cell(row=row_num, column=1, value=instance_id)
        worksheet.cell(row=row_num, column=2, value=instanceName)
        worksheet.cell(row=row_num, column=3, value=eventname)
        worksheet.cell(row=row_num, column=4, value=username)
        worksheet.cell(row=row_num, column=5, value=access_key)
        worksheet.cell(row=row_num, column=6, value=eventtime.strftime('%Y-%m-%d %H:%M:%S'))

        row_num +=1

workbook.save('EC2_Cloudtrail.xlsx')